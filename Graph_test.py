'''
1.从excel数据建立带权边列表
2.画出有向带权图Map
2'.画图(可不做)
3.dijkstra算法 算出两点间最短路线path 并计算距离distance
4.输入distance,计算价格
5.通过GUI输入站点,输出价格
'''

##从excel数据建立带权边列表
from openpyxl import load_workbook

#加载数据
wb=load_workbook('subway_db.xlsx')
ws = wb[wb.sheetnames[0]]

#终止函数
def check_None(r,c):
    if ws.cell(row = r, column = c).value == None and (
                ws.cell(row = r+1, column = c).value == None):
        return 0
    else:return 1
edges=[]
ri=1
ci=1
while check_None(ri,ci):
    start = ws.cell(row = ri , column = ci).value
    end = ws.cell(row = ri , column = ci+1).value
    weight = ws.cell(row = ri , column = ci+2).value
    if start != end and weight != None:
        edges.append((start,end,weight))
    ri+=1
#print(edges)
##

##画出有向带权图Map
import networkx as nx

#新建图
Map = nx.Graph()

#增加有权边
Map.add_weighted_edges_from(edges)

'''
#画图
pos=nx.spectral_layout(Map)
elarge=[(u,v) for (u,v,d) in Map.edges(data=True) if d['weight'] >1692]
esmall=[(u,v) for (u,v,d) in Map.edges(data=True) if d['weight'] <=1692]
nx.draw_networkx_edges(Map,pos,width=1)
nx.draw_networkx_nodes(Map,pos,node_size=5,node_shape='*')

plt.show()
plt.savefig("Map.png")
'''

##dijkstra算法 算出两点间最短距离
def dijkstra(Map, start, end):
    Map_0 = Map
    dlength = {}
    dupdate = {}
    for node in Map_0.nodes():
        dlength[node] = float('inf')       #初始距离为无穷大--路线长度字典dlength
        dupdate[node] = 'none'          #更新列表初始化
    dlength[start] = 0                       #起点到起点距离为0
    mid = start                           #将初始点置于操作中
    #print('最小距离字典:')#print(dlength)
    #print('最小距离点:')#print(min(dlength, key=dlength.get))
#搜索所有到终点的路径
    while mid != end:
        #获得最小值对应的键#print('--处理节点:')#print(mid)
        mid = min(dlength, key=dlength.get)   #取路线长度字典中值最小节点
        dlength_mid = dlength[mid]            #距离值赋值给变量
        del dlength[mid]                      #释放节点,以遍历
        #print('--节点路径:')#print(Map_0.edges(mid))
        #print('--中节点到各节点距离:')#print(Map_0[mid])
        for mid, update in Map_0.edges(mid):
            if update in dlength:
                new_length = dlength_mid + Map_0[mid][update]['weight']     #更新到该节点权重
                if new_length < dlength[update]:
                    dlength[update] = new_length
                    dupdate[update] = mid
    road = (end,)
    last = end
    b=0
    while last != start:
        nxt = dupdate[last]
        road += (nxt,)
        last = nxt
    distance = 0
    print(road)
    return road
def distance(Map,road):
    distance = 0
    for i in range(len(road)-1):
        starti = road[i]
        endi = road[i+1]
        weight = Map.get_edge_data(starti,endi)
        distance += weight['weight']
    path = ''
    for i in range(len(road)):
        path += road[len(road)-i-1]
        if i != len(road)-1:
            path += ','
    print(distance)
    if road[0] =='':
        return 0
    elif len(road) ==1:
        return 1
    else: return distance
    #return path
#distance = dijkstra(Map,'五道口','六道口')
#print(distance)
##计算价格
def price(distance):
    d=distance
    if d != 1:
        p = 3
        if d>32000:
            p=6+int((d-32000)/20000)
        elif d>22000:
            p=6
        elif d>12000:
            p=5
        elif d >6000:
            p=4
        if d:return p
        else:return 'empty'
    if d ==1:
        return 'same'
#price=price(distance)
#print(price)

##
##将输入输出GUI
import tkinter as tk
from tkinter import messagebox
import pickle

#窗口/标题
window = tk.Tk()
window.title('北京地铁计价系统')
window.geometry('320x500')
window.resizable(width=False,height=False)

#画布/背景
canvas = tk.Canvas(window, height=320, width=500)
bg_file = tk.PhotoImage(file='Map.png')
bg = canvas.create_image(0,0, anchor ='nw',image=bg_file)
canvas.pack(side='top',anchor='nw')#显示

# user information
tk.Label(window, text='起点: ').place(x=20,y= 250)
tk.Label(window, text='终点: ').place(x=20, y= 300)

text = tk.Text(window, height = 5,width = 27,font = ('Arial',14),x=100,y=350)
txt=('地铁票价: 6公里（含）内3元；6公里至12公里（含）4元；'
     '12公里至22公里（含）5元；22公里至32公里（含）6元；'
     '32公里以上部分，每增加1元可乘坐20公里.')
text.insert(tk.INSERT, txt)
text.pack(anchor ='n')


input_start = tk.StringVar()
start = tk.Entry(window, textvariable=input_start)
start.place(x=60, y=250)
input_end = tk.StringVar()
end = tk.Entry(window, textvariable=input_end)
end.place(x=60, y=300)

def look_up():
    map_start = input_start.get()
    map_end = input_end.get()
    try:
        road = dijkstra(Map,map_start,map_end)
        distance0 = distance(Map,road)
        p = price(distance0)
        print(p)
        if p =='empty':
            tk.messagebox.showerror(title='输入为空', message='请填写站点名')
        elif p == 'same':
            tk.messagebox.showerror(title='起点终点重复', message='票价是0元')
        else:
            message_0='票价是'+str(p)+'元'+'路径:'
            print(message_0)
            for station in road :
                message_0 += station
                if station != input_start:
                    message_0 += '\n'
                    message_0 += '↑'
                    message_0 += '\n'
            tk.messagebox.showinfo(title='路径搜索完毕', message=message_0)
    except KeyError:
        tk.messagebox.showerror(title='错误', message='找不到对应的站点信息')
    except ValueError:
        tk.messagebox.showerror(title='错误', message='找不到对应的站点信息')


bt_confirm = tk.Button(window, text='搜索',command = look_up)
bt_confirm.place(x=264, y=240,height= 100,width = 50)

window.mainloop()
